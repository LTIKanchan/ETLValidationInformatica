from pathlib import Path

import cx_Oracle
import openpyxl
import pandas as pd
import TestCase.DuplicateCountCheckTestCase as td
import TestCase.DataValidationTestCase as tv
import TestCase.ReportGenerator as tr

def executetest(excelpath, wordfile):
    try:
        print("****** Execution Started of Python Data verification script ******")
        cx_Oracle.init_oracle_client(lib_dir=r"C:\Oracle\instantclient_19_11")
        scenarioname = "None"
        workbook = openpyxl.load_workbook(excelpath, data_only=True)
        ws = workbook.sheetnames
        getsheet = ws[0]
        exceldata = pd.read_excel(excelpath, sheet_name=getsheet)
        if exceldata.iloc[0][1] == 'Yes' and exceldata.iloc[1][1] == 'Yes':
            print("* As per Verification flag , Started executing Both Data Validation and Duplicate Count Check *")
            td.checkduplicatedatacount(excelpath, wordfile)
            tv.comparedata(excelpath, wordfile)
            scenarioname = "Both"
        elif exceldata.iloc[0][1] == 'Yes' and exceldata.iloc[1][1] == 'No':
            print("* As per Verification flag , Started executing Duplicate Count Check Only *")
            td.checkduplicatedatacount(excelpath,wordfile)
            scenarioname = "Duplicate"
        elif exceldata.iloc[0][1] == 'No' and exceldata.iloc[1][1] == 'Yes':
            print("* As per Verification flag , Started executing Data Validation Only *")
            tv.comparedata(excelpath,wordfile)
            scenarioname = "DataValidation"
        elif exceldata.iloc[0][1] == 'No' and exceldata.iloc[1][1] == 'No':
            print("****** Both Flags in Pre Validation Check sheet is Set as 'No'.Hence execution is stopped.******")
        else:
            print("****** Invalid Input in Pre Validation Check Sheet ******")

        tr.getexecutionresult(scenarioname)
        print("****** Execution of Python Data verification script Completed Successfully ******")
    except Exception as ex:
        raise ex

def main():
    excelpath = str(Path(__file__).parent.parent) + "\\ETLDataValidation\\TestData\\TestDataFile.xlsx"
    wordfile = str(Path(__file__).parent.parent) + "\\ETLDataValidation\\TestData\\config.txt"
    executetest(excelpath, wordfile)


if __name__ == "__main__":
    main()
