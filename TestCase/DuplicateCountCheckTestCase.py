from pathlib import Path
import cx_Oracle
import pandas as pd
import sqlalchemy
import os
import openpyxl
from datetime import datetime
import pyodbc as pyodbc  #easily connect Python applications to data sources with an ODBC driver
from ConnectionInputs.GetCred import *  # Here we have taken .Getcredential file from ConnectionInput folder


def checkduplicatedatacount(excelpath, wordfile):  #here we have created function as checkduplicatedatacount
    #we have taken parameter testdata file and credential file
    cx_Oracle.init_oracle_client(lib_dir=r"C:\Oracle\instantclient_19_11")
    cflag='True'  # If flag is true, connection established to oracle client
    theexception=""
    try:  #it is used for critical statements
        print("$$$$$ Duplicate Count Test Case Execution Started $$$$$")
        now = datetime.now()  #used to take current date and time
        todaysdate = now.strftime("%d%m%Y %H%M")  #strftime is used to represent date& time in string format
        resultpath = str(Path(__file__).parent.parent) + "\\TestExecutionResults\\DuplicateCountResult_" + str(
            todaysdate)  #we have taken result path file with todays date in str format
        workbook = openpyxl.load_workbook(excelpath, data_only=True)  #it will search for excel file and data in that file
        ws = workbook.sheetnames  # will return name of worksheet from the workbook
        getsheet = ws[1]  # we have to take sheet number 2 (duplicate count sheet)
        worksheet = workbook[getsheet]   # it will take the number 2 worksheet
        oracle_usernamepass = []  #for oracle username & password
        exceldata = pd.read_excel(excelpath, sheet_name=getsheet)  # will read the excel path sheets
        for i in range(len(exceldata['Database Type'])):  # will check database type column from sheet (count -5)
            cflag = 'True'
            testcaseno = exceldata.iloc[i][0]  # it will read data from 0 to 4th location)
            if exceldata.iloc[i][3] == 'Yes':  # this if condition will check the 3rd location column
                # from duplicatecount sheet (validation condtion )
                oracle_hostname = exceldata.iloc[i][4]
                oracle_portno = exceldata.iloc[i][5]
                oracle_service = exceldata.iloc[i][6]
                oracle_usernamepass = readtext(wordfile, oracle_service)
                #readtext will check wordfile,servernameintextfile for username & password:
                oracle_username = oracle_usernamepass[0]
                oracle_password = oracle_usernamepass[1]
                # print(oracle_username+" "+oracle_password)

                if exceldata.iloc[i][2].upper() == 'SQL':  # this block will connect to SQL DB
                    sqldbserver = exceldata.iloc[i][4]
                    sqldbname = exceldata.iloc[i][6]
                    sql_engine = sqlalchemy.create_engine(
                    'mssql://' + sqldbserver + '/' + sqldbname + '?driver=ODBC+Driver+13+for+SQL+Server?Trusted_Connection=yes')
                    query = exceldata.iloc[i][1]
                    sourcedata = pd.read_sql(query, con=sql_engine)

                elif exceldata.iloc[i][2].upper() == 'ORACLE':  # this block will connect to oracle DB
                    try:
                        #cflag = 'True'
                        ordsn_tns = cx_Oracle.makedsn(oracle_hostname, oracle_portno, service_name=oracle_service)
                        oracle_engine = cx_Oracle.connect(user=oracle_username, password=oracle_password, dsn=ordsn_tns)
                        query = exceldata.iloc[i][1]
                        sourcedata = pd.read_sql(query, con=oracle_engine)

                    except Exception as ex:
                        cflag='false'
                        theexception=ex

                if not os.path.exists(resultpath):  # is used to check whether the specified path exists or not
                    os.mkdir(resultpath)  #mkdir() method in Python is used to create a directory named path with the specified numeric mode.
                    # This method raise FileExistsError if the directory to be created already exists.

                for k in range(i, i + 1):
                    if(cflag=='True'):
                        worksheet.cell(k + 2, 8).value = sourcedata.shape[0]
                        diffresult = worksheet.cell(k + 2, 8).value  #Number of Records Having Duplicate Count
                        if diffresult > 0 and diffresult <= 100:  # 1st condition
                            resultfile = resultpath + "\\" + str(testcaseno) + "ExecutionResult.xlsx"
                            for j in range(len(sourcedata.columns)):
                                limitdata = sourcedata.head(diffresult)
                                limitdata.to_excel(resultfile, index=False)
                                worksheet.cell(k + 2, 9).value = resultfile
                                worksheet.cell(k + 2, 10).value = "Fail"
                                worksheet.cell(k + 2, 11).value = "Duplicate Count "+str(diffresult)
                            print("Duplicate Count Check " + str(testcaseno) + " is Failed....")
                            workbook.save(excelpath)

                        elif diffresult >100:  # 2nd condition check
                            resultfile = resultpath + "\\" + str(testcaseno) + "ExecutionResult.xlsx"
                            for j in range(len(sourcedata.columns)):
                                limitdata = sourcedata.head(500)
                                limitdata.to_excel(resultfile, index=False)
                                worksheet.cell(k + 2, 9).value = resultfile
                                worksheet.cell(k + 2, 10).value = "Fail"
                                worksheet.cell(k + 2, 11).value = "Duplicate Count "+str(diffresult)
                            print("Duplicate Count Check " + str(testcaseno) + " is Failed....")
                            workbook.save(excelpath)

                        elif diffresult == 0:  #3rd condition
                            worksheet.cell(k + 2, 10).value = "Pass"
                            worksheet.cell(k + 2, 11).value = "No Duplicate Count found"
                            print("Duplicate Count Check " + str(testcaseno) + " is Passed....")
                        workbook.save(excelpath)

                    else:
                        worksheet.cell(k + 2, 10).value = "Fail"
                        worksheet.cell(k + 2, 11).value = str(theexception)
                        worksheet.cell(k + 2, 8).value = "0"
                        worksheet.cell(k + 2, 9).value = "NA"
                        print(str(testcaseno)+" Test case failed due to exception "+str(theexception))
                    workbook.save(excelpath)

            elif exceldata.iloc[i][3] == 'No':  # this if condition will check the 3rd location column
                # from duplicatecount sheet (validation condtion )
                worksheet.cell(i + 2, 8).value = "0"
                worksheet.cell(k + 2, 9).value = "NA"
                worksheet.cell(i + 2, 10).value = "No Run"
                worksheet.cell(i + 2, 11).value = "Test Case Execution Skipped"
                print("Duplicate Count Check " + str(testcaseno) + " is Not Run....")
                workbook.save(excelpath)
        workbook.save(excelpath)
        print("$$$$$ Data Duplicate Count Test Case Execution Completed $$$$$")
        print("")
    except Exception as ex:
        raise ex


def main():
    excelpath = "C:\\Users\\bshinde\\PycharmProjects\\ETLValidationBshinde\\TestData\\TestDataFile.xlsx"
    #excelpath = str(Path(__file__).parent.parent) + "\\TestData\\TestDataFile.xlsx"  # path of test data file
    wordfile = str(Path(__file__).parent.parent) + "\\TestData\\config.txt"  # path of username & password Encrypted file
    checkduplicatedatacount(excelpath, wordfile)


if __name__ == "__main__":
    main()
