from pathlib import Path
import cx_Oracle
import pandas as pd
import sqlalchemy
import os
import openpyxl
from datetime import datetime
import pyodbc as pyodbc
from ConnectionInputs.GetCred import *  # #Here we have taken all the data from GetCred file from ConnectionInput folder


def checkduplicatedatacount(excelpath, wordfile):
    try:
        print("$$$$$ Duplicate Count Test Case Execution Started $$$$$")
        now = datetime.now()
        todaysdate = now.strftime("%d%m%Y %H%M")
        resultpath = str(Path(__file__).parent.parent) + "\\TestExecutionResults\\DuplicateCountResult_" + str(
            todaysdate)
        workbook = openpyxl.load_workbook(excelpath, data_only=True)
        ws = workbook.sheetnames
        getsheet = ws[1]
        worksheet = workbook[getsheet]
        oracle_usernamepass = []
        exceldata = pd.read_excel(excelpath, sheet_name=getsheet)
        for i in range(len(exceldata['Database Type'])):
            try:
                testcaseno = exceldata.iloc[i][0]
                if exceldata.iloc[i][3] == 'Yes':
                    oracle_hostname = exceldata.iloc[i][4]
                    oracle_portno = exceldata.iloc[i][5]
                    oracle_service = exceldata.iloc[i][6]
                    oracle_usernamepass = readtext(wordfile, oracle_service)
                    oracle_username = oracle_usernamepass[0]
                    oracle_password = oracle_usernamepass[1]
                # print(oracle_username+" "+oracle_password)

                if exceldata.iloc[i][2].upper() == 'SQL':
                    sqldbserver = exceldata.iloc[i][4]
                    sqldbname = exceldata.iloc[i][6]
                    sql_engine = sqlalchemy.create_engine(
                    'mssql://' + sqldbserver + '/' + sqldbname + '?driver=ODBC+Driver+13+for+SQL+Server?Trusted_Connection=yes')
                    query = exceldata.iloc[i][1]
                    sourcedata = pd.read_sql(query, con=sql_engine)

                elif exceldata.iloc[i][2].upper() == 'ORACLE':
                    ordsn_tns = cx_Oracle.makedsn(oracle_hostname, oracle_portno, service_name=oracle_service)
                    oracle_engine = cx_Oracle.connect(user=oracle_username, password=oracle_password, dsn=ordsn_tns)
                    #print("@@@@ Connection Successful @@@@")
                    query = exceldata.iloc[i][1]
                    sourcedata = pd.read_sql(query, con=oracle_engine)
                        # print(sourcedata)

                if not os.path.exists(resultpath):
                    os.mkdir(resultpath)

                    for k in range(i, i + 1):
                        worksheet.cell(k + 2, 10).value = sourcedata.shape[0]
                        diffresult = worksheet.cell(k + 2, 10).value
                        if diffresult >0 and diffresult <=100  :
                            resultfile = resultpath + "\\" + str(testcaseno) + "ExecutionResult.xlsx"
                            for j in range(len(sourcedata.columns)):
                                limitdata = sourcedata.head(diffresult)
                                limitdata.to_excel(resultfile, index=False)
                                worksheet.cell(k + 2, 11).value = resultfile
                                worksheet.cell(k + 2, 12).value = "Fail"
                                print("Duplicate Count Check " + str(testcaseno) + " is Failed....")
                        elif diffresult > 100:
                            resultfile = resultpath + "\\" + str(testcaseno) + "ExecutionResult.xlsx"
                            for j in range(len(sourcedata.columns)):
                                limitdata = sourcedata.head(100)
                                limitdata.to_excel(resultfile, index=False)
                                worksheet.cell(k + 2, 11).value = resultfile
                                worksheet.cell(k + 2, 12).value = "Fail"
                            print("Duplicate Count Check " + str(testcaseno) + " is Failed....")
                        elif diffresult == 0:
                            worksheet.cell(k + 2, 12).value = "Pass"
                        print("Duplicate Count Check " + str(testcaseno) + " is Passed....")
                        workbook.save(excelpath)
                elif exceldata.iloc[i][3] == 'No':
                    worksheet.cell(i + 2, 12).value = "No Run"
                    print("Duplicate Count Check " + str(testcaseno) + " is Not Run....")
            except Exception as er:
                pass
                workbook.save(excelpath)
                print("$$$$$ Data Duplicate Count Test Case Execution Completed $$$$$")
                print("")
    except Exception as ex:
        raise ex


def main():
    excelpath = str(Path(__file__).parent.parent) + "\\Test Data\\TestDataFile.xlsx"
    wordfile = str(Path(__file__).parent.parent) + "\\Test Data\\config.txt"
    checkduplicatedatacount(excelpath, wordfile)


if __name__ == "__main__":
    main()
