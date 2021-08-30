from pathlib import Path
import cx_Oracle  #It enables access to oracle DB
import pandas as pd  #pd is alias name foe pandas. It is used represent data in series,dataframe,panel
import sqlalchemy  #it is a library that faciliates comm'n betwn python prog & db's
import os  #it provides functions for interacting with operating systems
import openpyxl  #it is python library usd to read & write excel files
from datetime import datetime  # to used date and time in py program
import numpy as np  #it is used for multidimensional arrays(allows you to bind the import to the local variable name of your choice)
from ConnectionInputs.GetCred import *  # Here we have taken .Getcredential file from ConnectionInput folder
#this code change for test different branch affect and code change done for kpawar branch not working

def comparedata(excelpath,wordfile):  # here we have created function as comparedata, And parameter taken as textdata file & cred file
    cx_Oracle.init_oracle_client(lib_dir=r"C:\Oracle\instantclient_19_11")
    cflag = 'True'  # If flag is true, connection established to oracle client
    theexception = ""
    try:  #it is used for critical statements
        print("@@@@@ Data Validation Test Case Execution Started @@@@@")
        sqldbserver = ''  # SQL DB server
        sqldbname = ''  # SQL DB name
        now = datetime.now()  #used to take current date and time
        todaysdate = now.strftime("%d%m%Y %H%M")  #strftime is used to represent date& time in string format
        resultpath = str(Path(__file__).parent.parent) + "\\TestExecutionResults\\DataValidationResult_" + str(
            todaysdate)  #we have taken result path file with todays date in str format
        workbook = openpyxl.load_workbook(excelpath, data_only=True)
        #it will load/read testdata file path
        ws = workbook.sheetnames  # will take worksheet from the workbook(testdata)
        getsheet = ws[2]  # we have take sheet number 2 (i.e.data validation sheet)
        worksheet = workbook[getsheet]  # it will take the number 2 worksheet
        exdata = pd.read_excel(excelpath, sheet_name=getsheet)  # will read testdata file,data validation sheet
        for i in range(len(exdata['Source Type'])):  #it will check source type column in datavalidation sheet (count is 8)
            testcaseno = exdata.iloc[i][0]  # it will read data from 0th to 7th position
            if exdata.iloc[i][5].upper() == 'YES':  # this if condition will check the 5th location column
                # from datavalidation sheet (data validation condtion )
                oracle_hostname = exdata.iloc[i][6]
                oracle_portno = exdata.iloc[i][7]
                oracle_service = exdata.iloc[i][8]
                oracle_usernamepass = readtext(wordfile, oracle_service)  #it will check the readtext function from getcred file
                oracle_username = oracle_usernamepass[0]
                oracle_password = oracle_usernamepass[1]

                if exdata.iloc[i][2].upper() == 'SQL':  # this block will connect to SQL DB , if we are executing query from SQl server
                    sql_engine = sqlalchemy.create_engine(
                        'mssql://' + sqldbserver + '/' + sqldbname + '?driver=ODBC+Driver+13+for+SQL+Server?Trusted_Connection=yes')
                    query = exdata.iloc[i][1]
                    sourcedata = pd.read_sql(query, con=sql_engine)

                elif exdata.iloc[i][2].upper() == 'ORACLE':  # if it is not sql server , it will check for oracle DB query
                    try:
                        cflag = 'True'  # if for oracle connection flag is true, for connecting to required db
                        # we have to pass all the connection details
                        ordsn_tns = cx_Oracle.makedsn(oracle_hostname, oracle_portno, service_name=oracle_service)
                        oracle_engine = cx_Oracle.connect(user=oracle_username, password=oracle_password, dsn=ordsn_tns)
                        query = exdata.iloc[i][1]
                        sourcedata = pd.read_sql(query, con=oracle_engine)
                    except Exception as ex:  #it is used to throw exception , if try block not succeed to execute
                        cflag = 'false'
                        theexception = ex

                if exdata.iloc[i][4].upper() == 'SQL':  #target db connection, if it is sql DB
                    sql_engine = sqlalchemy.create_engine(
                        'mssql://' + sqldbserver + '/' + sqldbname + '?driver=ODBC+Driver+13+for+SQL+Server?Trusted_Connection=yes')
                    query = exdata.iloc[i][3]
                    targetdata = pd.read_sql(query, con=sql_engine)

                elif exdata.iloc[i][4].upper() == 'ORACLE': #target connection, if it is source DB
                    try:
                        #cflag = 'True'  # if for oracle connection flag is true, for connecting to required db
                        # we have to pass all the connection details
                        ordsn_tns = cx_Oracle.makedsn(oracle_hostname, oracle_portno, service_name=oracle_service)
                        oracle_engine = cx_Oracle.connect(user=oracle_username, password=oracle_password, dsn=ordsn_tns)
                        query = exdata.iloc[i][3]
                        targetdata = pd.read_sql(query, con=oracle_engine)
                    except Exception as ex:  #it is used to throw exception , if try block not succeed to execute
                        cflag = 'false'
                        theexception = ex

                if not os.path.exists(resultpath):  # is used to check whether the specified path exists or not
                    os.mkdir(resultpath)  #mkdir() method in Python is used to create a directory named path with the specified numeric mode.
                    # This method raise FileExistsError if the directory to be created already exists.

                for k in range(i + 1, i + 2):
                    if (cflag == 'True'):  #comparision operatio if cflag == true then below statements execute
                        Actualtargetdata = targetdata.replace(np.nan, 'No Value')  #replace function will replace blank value with string no value
                        Actualsourcedata = sourcedata.replace(np.nan, 'No Value')
                        diffresulttargetdata = pd.DataFrame(Actualtargetdata)  # This will covert Actualtargetdata in tabular format
                        diffresultsourcedata = pd.DataFrame(Actualsourcedata)  # This will covert Actualsourcedata in tabular format
                        diffresult = pd.DataFrame()  # empty dataframe object

                        if Actualsourcedata.shape[0] == Actualtargetdata.shape[0]:  # shape[0] means first dimension of array
                            assert_equal = np.array_equal(diffresulttargetdata, diffresultsourcedata)
                            if assert_equal:  # will check  check if two variables are equal,
                                print("Data Validation " + str(testcaseno) + " is Passed....")
                                for k in range(i + 1, i + 2):  # for pass test case
                                    worksheet.cell(k + 1, 10).value = sourcedata.shape[0]  # will gie you source count
                                    worksheet.cell(k + 1, 11).value = targetdata.shape[0]  # will gie you target count
                                    worksheet.cell(k + 1, 12).value = diffresult.shape[0]  # will give you count 0 for match columns
                                    worksheet.cell(k + 1, 13).value = "-"  # columns match , so will give you -
                                    worksheet.cell(k + 1, 14).value = "NA"  # for pass test case it will give NA
                                    worksheet.cell(k + 1, 15).value = "Pass"  # test case status
                                    worksheet.cell(k + 1, 16).value = "Data Validated Successfully"  # remark for pass test case as - 'no remark'
                                    workbook.save(excelpath)
                                #all changes not updated
                            else:
                                diffresult = diffresultsourcedata  #mismatch count of target
                                print("Data Validation " + str(testcaseno) + " is Failed....")

                                if diffresult.shape[0] > 0 and diffresult.shape[0] <= 500:
                                # will check the data should > 0 and <=100 (dump 100 records)
                                    mismatchcols = []  #empty list
                                    for j in range(len(Actualsourcedata.columns)):
                                        if not (Actualsourcedata[Actualsourcedata.columns[j]]).equals(
                                                Actualtargetdata[Actualtargetdata.columns[j]]):
                                            mismatchcols.append(Actualsourcedata.columns[j])
                                    resultfile = resultpath + "\\" + str(testcaseno) + "_ExecutionResult.xlsx"
                                    with pd.ExcelWriter(resultfile) as writer:
                                        limittargetdata = diffresulttargetdata.head(diffresult)
                                        limitsourcedata = diffresultsourcedata.head(diffresult)
                                        limittargetdata.to_excel(writer, sheet_name='Targetdata')
                                        limitsourcedata.to_excel(writer, sheet_name='Sourcedata')
                                        workbook.save(excelpath)

                                else:
                                    mismatchcols = []
                                    for j in range(len(Actualsourcedata.columns)):
                                        if not (Actualsourcedata[Actualsourcedata.columns[j]]).equals(
                                                Actualtargetdata[Actualtargetdata.columns[j]]):
                                            mismatchcols.append(Actualsourcedata.columns[j])
                                    resultfile = resultpath + "\\" + str(testcaseno) + "_ExecutionResult.xlsx"
                                    with pd.ExcelWriter(resultfile) as writer:  # used to write text,num,string data
                                        limittargetdata = diffresulttargetdata.head(100)  # head used to get first 100 rows
                                        limitsourcedata = diffresultsourcedata.head(100)
                                        limittargetdata.to_excel(writer, sheet_name='Targetdata')
                                        limitsourcedata.to_excel(writer, sheet_name='Sourcedata')
                                        workbook.save(excelpath)

                                for k in range(i + 1, i + 2):  # for fail test case
                                    worksheet.cell(k + 1, 10).value = sourcedata.shape[0]  # will gie you source count
                                    worksheet.cell(k + 1, 11).value = targetdata.shape[0]  # will gie you target count
                                    worksheet.cell(k + 1, 12).value = diffresult.shape[0]  # will give you mismatch columns count
                                    worksheet.cell(k + 1, 13).value = str(', '.join(mismatchcols))  # will give mismatch cols names
                                    worksheet.cell(k + 1, 14).value = resultfile  # for fail test cas, file path
                                    worksheet.cell(k + 1, 15).value = "Fail"  # test case status
                                    worksheet.cell(k + 1, 16).value = "Mismatch data found"  # remark for fail test case as - 'no remark'
                                    workbook.save(excelpath)
                        else:
                            print("Data Validation " + str(
                                testcaseno) + " Failed due to source and target table count mismatched..")
                            resultfile = resultpath + "\\" + str(testcaseno) + "_ExecutionResult.xlsx"
                            for k in range(i + 1, i + 2):
                                worksheet.cell(k + 1, 10).value = sourcedata.shape[0]
                                worksheet.cell(k + 1, 11).value = targetdata.shape[0]
                                worksheet.cell(k + 1, 12).value = diffresulttargetdata.shape[0]
                                worksheet.cell(k + 1, 13).value = "Source table and Target table count mismatched."
                                worksheet.cell(k + 1, 14).value = resultfile
                                worksheet.cell(k + 1, 15).value = "Fail"
                                worksheet.cell(k + 1, 16).value = "Source table and Target table count mismatched."
                                workbook.save(excelpath)
                            with pd.ExcelWriter(resultfile) as writer:
                                limittargetdata = diffresulttargetdata.head(500)
                                limitsourcedata = diffresultsourcedata.head(500)
                                limittargetdata.to_excel(writer, sheet_name='Targetdata')
                                limitsourcedata.to_excel(writer, sheet_name='Sourcedata')
                                workbook.save(excelpath)

                    else:
                        worksheet.cell(k + 1, 10).value = '-'
                        worksheet.cell(k + 1, 11).value = '-'
                        worksheet.cell(k + 1, 12).value = '-'
                        worksheet.cell(k + 1, 13).value = "-"
                        worksheet.cell(k + 1, 14).value = "NA"
                        worksheet.cell(k + 1, 15).value = "Fail"
                        worksheet.cell(k + 1, 16).value = str(theexception) # technical error excepion will be print
                        print(str(testcaseno) + " Test case failed due to exception " + str(theexception))
                    workbook.save(excelpath)

            elif exdata.iloc[i][5] == 'No':  # if 5th location condition in no (data validation condition)
                worksheet.cell(k + 1, 10).value = '-'
                worksheet.cell(k + 1, 11).value = '-'
                worksheet.cell(k + 1, 12).value = '-'
                worksheet.cell(k + 1, 13).value = "-"
                worksheet.cell(i + 2, 14).value = "NA"
                worksheet.cell(i + 2, 15).value = "No Run"
                worksheet.cell(i + 2, 16).value = "Test Case Execution Skipped"
                print("Data Validation " + str(testcaseno) + " is Not Run....")
                workbook.save(excelpath)
        workbook.save(excelpath)
        print("@@@@@ Data Validation Test Case Execution Completed @@@@@")
        print("")
    except Exception as ex:
        raise ex  #The raise keyword is used to raise an exception. You can define what kind of error to raise

def main():
    #excelpath = "C:\\Users\\bshinde\\PycharmProjects\\ETLValidationBshinde\\TestData\\TestDataFile.xlsx"
    excelpath = str(Path(__file__).parent.parent) + "\\TestData\\TestDataFileAll.xlsx"  # path of test data file
    wordfile = str(Path(__file__).parent.parent) + "\\TestData\\config.txt"  # path of username & password Encrypted file
    comparedata(excelpath, wordfile)

if __name__ == "__main__":
    main()
