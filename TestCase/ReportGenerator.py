from datetime import datetime
import os
from pathlib import Path
import openpyxl
import pandas as pd
import xlsxwriter  #is a Python module that can be used to write text, numbers & formulas
import matplotlib.pyplot as plt  #matplotlib. pyplot is a collection of functions,pyplot function makes some change to a figure: e.g., creates a figure, creates a plotting area in a figure,
from openpyxl.styles import *
from openpyxl.utils.dataframe import dataframe_to_rows  #passing the value to worksheet
from pandas.plotting import table


def getexecutionresult(scenarioname):

    now = datetime.now()
    datafilename = str(Path(__file__).parent.parent) + "\\TestData\\TestDataFile.xlsx"
    todaysdate = now.strftime("%d%m%Y %H%M")
    workbook = openpyxl.load_workbook(datafilename, data_only=True)   #it will load test data file in workbook
    ws = workbook.sheetnames
    filename = str(Path(__file__).parent.parent) + "\\ExecutionReports"
    resultfile = filename + "\\ExecutionReport_"+str(todaysdate)+".xlsx"
    if not os.path.exists(filename):  #it will check ExecutionReport file exist or not?
        os.mkdir(filename)  #if not present will create the directory

    wbk = xlsxwriter.Workbook(resultfile)   #Will write the ExecutionReport file in to Wbk
    #Alignment =wbk.add_format({'text_wrap': True})
    wbk.close()  #to close workbook file

    for i in range(len(ws)):  # will return num of items in worksheet
        passtestcases = 0
        failtestcases = 0
        noruntestcases = 0
        flag = "No"
        if scenarioname == "Both":
            if ws[i] == 'Duplicate Count':
                print("##### Duplicate count execution report generating #####")
                ExecutionResult = pd.read_excel(datafilename, sheet_name=ws[i])  #will read testdata file
                testcasename = ExecutionResult['Test Case Name']
                resultfilename = ExecutionResult['Result File Path']
                testcasestatus = ExecutionResult['Test Case Status']
                remarkoftestcase = ExecutionResult['Remarks']
                number= ExecutionResult['Sr.No']
                resultduplicate = pd.concat([number, testcasename, testcasestatus, resultfilename, remarkoftestcase], axis=1, ignore_index=False)
                dfdup = pd.DataFrame(resultduplicate)
                flag = "Yes"

            if ws[i] == 'Data Validation':
                print("##### Data Validation execution report generating #####")
                ExecutionResult = pd.read_excel(datafilename, sheet_name=ws[i])
                testcasename = ExecutionResult['Test Case Name']
                resultfilename = ExecutionResult['File Path']
                testcasestatus = ExecutionResult['Test Case Status']
                remarkoftestcase = ExecutionResult['Remark']
                number=ExecutionResult['Sr.No']
                resultvalidation = pd.concat([number,testcasename, testcasestatus, resultfilename, remarkoftestcase], axis=1, ignore_index=False)
                dfval = pd.DataFrame(resultvalidation)
                dfval.set_index(number)
                flag = "Yes"
        if scenarioname == "Duplicate":
            if ws[i] == 'Duplicate Count':
                print("##### Duplicate count execution report generating #####")
                ExecutionResult = pd.read_excel(datafilename, sheet_name=ws[i])
                testcasename = ExecutionResult['Test Case Name']
                resultfilename = ExecutionResult['Result File Path']
                testcasestatus = ExecutionResult['Test Case Status']
                remarkoftestcase = ExecutionResult['Remarks']
                number = ExecutionResult.shape[0]
                resultduplicate = pd.concat([number,testcasename, testcasestatus, resultfilename, remarkoftestcase], axis=1, ignore_index=False)
                dfdup = pd.DataFrame(resultduplicate)
                flag = "Yes"

        if scenarioname == "DataValidation":
            if ws[i] == 'Data Validation':
                print("##### Data Validation execution report generating #####")
                ExecutionResult = pd.read_excel(datafilename, sheet_name=ws[i])
                testcasename = ExecutionResult['Test Case Name']
                resultfilename = ExecutionResult['Result File Path']
                testcasestatus = ExecutionResult['Test Case Status']
                remarkoftestcase = ExecutionResult['Remarks']
                resultvalidation = pd.concat([number,testcasename, testcasestatus, resultfilename, remarkoftestcase], axis=1, ignore_index=False)
                dfval = pd.DataFrame(resultvalidation)
                flag = "Yes"
        if scenarioname == "None":
            print("##### Execution Stopped. No Report is generated. #####")

        if flag == "Yes":
            totaltestcases = ExecutionResult.shape[0]
            statusname = ExecutionResult['Test Case Status']
            for j in range((statusname.shape[0])):
                if statusname[j] == 'Pass':
                    passtestcases = passtestcases + 1
                if statusname[j] == 'Fail':
                    failtestcases = failtestcases + 1
                if statusname[j] == 'No Run':
                    noruntestcases = noruntestcases + 1

            Tasks = [noruntestcases, failtestcases, passtestcases]
            my_labels = ['No Run', 'Fail', 'Pass']
            plt.figure(figsize=(13, 3))
            ax1 = plt.subplot(121, aspect='equal')
            plt.pie(Tasks, labels=my_labels, autopct='%1.1f%%')
            plt.title('Execution Result of ' + str(ws[i]))
            plt.axis('equal')
            raw_data = {'Test Execution Status': ['Pass', 'Fail', 'No Run', 'Total'],
                        'Count': [passtestcases, failtestcases, noruntestcases, totaltestcases]}
            df = pd.DataFrame(raw_data, columns=['Test Execution Status', 'Count'], index={1, 2, 3, 4})
            ax2 = plt.subplot(122)
            plt.axis('off')
            tbl = table(ax2, df, loc='center')
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(14)
            shtName = ws[i]
            fname = str(shtName) + str(todaysdate)
            plt.savefig(str(Path(__file__).parent.parent) + "\\Images\\" + fname + ".png")
            plt.close(fig=None)
            wbk = openpyxl.load_workbook(resultfile)
            if shtName.__contains__("Validation") or shtName.__contains__("Duplicate"):
                wp = wbk.create_sheet(shtName)
                imagepath = str(Path(__file__).parent.parent) + "\\Images\\" + fname + ".png"
                img = openpyxl.drawing.image.Image(imagepath)
                img.anchor = 'A1'
                wp.add_image(img)
                wbk.save(resultfile)
                if shtName.__contains__("Validation"):
                    rows = dataframe_to_rows(dfval, header=True)
                    for r_idx, row in enumerate(rows, 17):
                        for c_idx, mvalue in enumerate(row, 1):
                            wp.cell(row=r_idx, column=c_idx, value=mvalue)
                    wbk.save(resultfile)
                    wp.delete_rows(18)
                    wbk.save(resultfile)
                    wp['B17'].font = Font(bold=True)
                    wp['C17'].font = Font(bold=True)
                    wp['D17'].font = Font(bold=True)
                    wp['E17'].font= Font(bold=True)
                    wp['F17'].font = Font(bold=True)
                    wp.column_dimensions['A'].hidden = True
                    wbk.save(resultfile)
                    wp.column_dimensions['B'].width = 6
                    wp.column_dimensions['C'].width = 30
                    wp.column_dimensions['D'].width = 15
                    wp.column_dimensions['E'].width = 65
                    wp.column_dimensions['F'].width = 65
                    wbk.save(resultfile)
                    for i in range(18,wp.max_row+1):
                        for j in range(5,wp.max_column+1):
                            wrap_alignment = Alignment(wrap_text=True)
                            wp.cell(i,j).alignment = wrap_alignment
                    wbk.save(resultfile)

                if shtName.__contains__("Duplicate"):
                    rows = dataframe_to_rows(dfdup,header=True)
                    for r_idx, row in enumerate(rows, 17):
                        for c_idx, mvalue in enumerate(row, 1):
                            wp.cell(row=r_idx, column=c_idx, value=mvalue)
                    wbk.save(resultfile)
                    wp.delete_rows(18)
                    wbk.save(resultfile)
                    wp['B17'].font = Font(bold=True)
                    wp['C17'].font = Font(bold=True)
                    wp['D17'].font = Font(bold=True)
                    wp['E17'].font = Font(bold=True)
                    wp['F17'].font = Font(bold=True)
                    wp.column_dimensions['A'].hidden = True
                    wbk.save(resultfile)
                    wp.column_dimensions['B'].width = 6
                    wp.column_dimensions['C'].width = 30
                    wp.column_dimensions['D'].width = 15
                    wp.column_dimensions['E'].width = 65
                    wp.column_dimensions['F'].width = 65
                    wbk.save(resultfile)
                    for i in range(18,wp.max_row+1):
                        for j in range(5,wp.max_column+1):
                            wrap_alignment = Alignment(wrap_text=True)
                            wp.cell(i,j).alignment = wrap_alignment
                    wbk.save(resultfile)
                    for k in range(18,wp.max_row + 1):
                        for m in range(5,wp.max_column):
                            cellformat = Font(color='000000FF',underline='single')
                            wp.cell(k,m).font = cellformat
                            #wp.cell(k, m).style = link
                    wbk.save(resultfile)
                                #wp.cell(i,j).style= cellformat

            std = wbk.sheetnames
            for h in range(len(std)):
                if std[h] == "Sheet1":
                    ss = wbk['Sheet1']
                    wbk.remove(ss)
                wbk.save(resultfile)


def main():
    scenarioname = "Both"
    getexecutionresult(scenarioname)


if __name__ == "__main__":
    main()
